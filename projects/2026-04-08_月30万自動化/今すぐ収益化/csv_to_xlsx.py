#!/usr/bin/env python3
"""
全 CSV テンプレートを xlsx に変換するスクリプト。
Vol.2/5/6/10 の CSV を1ファイル xlsx に統合（複数シート構成）。

【使い方】
  python csv_to_xlsx.py

【出力】
  projects/.../C_テンプレ販売/dist/Vol2_SNSカレンダー.xlsx
  projects/.../C_テンプレ販売/dist/Vol5_確定申告準備.xlsx
  projects/.../C_テンプレ販売/dist/Vol6_クライアント管理DB.xlsx
  projects/.../C_テンプレ販売/dist/Vol10_売上ダッシュボード.xlsx
"""

import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent / "C_テンプレ販売"
DIST = ROOT / "dist"
DIST.mkdir(exist_ok=True)

HEADER_FILL = PatternFill(start_color="DAE9F8", end_color="DAE9F8", fill_type="solid")
HEADER_FONT = Font(bold=True)


def csv_to_sheet(wb, csv_path: Path, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    with csv_path.open(encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # 数式（=で始まる）は文字列として扱う（読者がコピーで使う想定）
                cell.value = val
                if row_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center")
    # 列幅を自動調整（簡易）
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = min(length, 50)
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 50))
    return ws


def build_workbook(vol_label: str, sheet_specs: list[tuple[str, Path]]):
    """sheet_specs: [(シート名, csv パス), ...]"""
    wb = openpyxl.Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)
    for name, path in sheet_specs:
        if not path.exists():
            print(f"[WARN] 見つかりません: {path}")
            continue
        csv_to_sheet(wb, path, name)
    out = DIST / f"{vol_label}.xlsx"
    wb.save(out)
    print(f"[OK] {out}")
    return out


# ─────────────────────────────────
# Vol.2 SNSカレンダー
# ─────────────────────────────────
build_workbook(
    "Vol2_SNSカレンダー",
    [
        ("月次カレンダー", ROOT / "vol2_sheet1_calendar.csv"),
        ("投稿テーマ50選", ROOT / "vol2_sheet2_themes.csv"),
    ],
)

# ─────────────────────────────────
# Vol.5 確定申告準備
# ─────────────────────────────────
build_workbook(
    "Vol5_確定申告準備",
    [
        ("月次収支入力", ROOT / "vol5_sheet1_transactions.csv"),
        ("月次集計", ROOT / "vol5_sheet2_monthly_summary.csv"),
        ("年間サマリ", ROOT / "vol5_sheet3_yearly_summary.csv"),
        ("経費科目チートシート", ROOT / "vol5_sheet4_expense_cheatsheet.csv"),
        ("青色vs白色", ROOT / "vol5_sheet5_blue_white_guide.csv"),
    ],
)

# ─────────────────────────────────
# Vol.6 クライアント管理DB
# ─────────────────────────────────
build_workbook(
    "Vol6_クライアント管理DB",
    [
        ("クライアント", ROOT / "vol6_db1_clients.csv"),
        ("案件", ROOT / "vol6_db2_jobs.csv"),
        ("請求", ROOT / "vol6_db3_invoices.csv"),
        ("契約", ROOT / "vol6_db4_contracts.csv"),
    ],
)

# ─────────────────────────────────
# Vol.10 売上ダッシュボード
# ─────────────────────────────────
build_workbook(
    "Vol10_売上ダッシュボード",
    [
        ("今月のサマリ", ROOT / "vol10_dashboard_sheet1_summary.csv"),
        ("取引", ROOT / "vol10_dashboard_sheet2_transactions.csv"),
        ("柱別月次", ROOT / "vol10_dashboard_sheet3_monthly_by_pillar.csv"),
        ("年間累計", ROOT / "vol10_dashboard_sheet4_yearly.csv"),
    ],
)

print()
print("=" * 60)
print(f"全 4 ファイルを {DIST} に出力しました。")
print("オーナーはこの xlsx を note にアップロードするだけで販売開始可能。")
