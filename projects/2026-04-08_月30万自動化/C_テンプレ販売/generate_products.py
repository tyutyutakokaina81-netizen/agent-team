"""
generate_products.py — 販売用テンプレートファイルを自動生成

生成物:
  dist/vol1_freelance_cashflow.xlsx   ← BOOTH に即アップロード可能
  dist/vol2_sns_calendar.xlsx         ← BOOTH に即アップロード可能

実行方法:
  pip install openpyxl
  python generate_products.py
"""

from pathlib import Path
from datetime import datetime

OUTPUT_DIR = Path(__file__).parent / "dist"
OUTPUT_DIR.mkdir(exist_ok=True)


# ─────────────────────────────────────────────────────────
# Vol.1: フリーランス収支管理スプレッドシート
# ─────────────────────────────────────────────────────────

def create_vol1():
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("[ERROR] pip install openpyxl")
        return None

    wb = openpyxl.Workbook()

    # ── スタイル定義 ──
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    accent_fill  = PatternFill("solid", fgColor="D6E4F0")
    total_fill   = PatternFill("solid", fgColor="FFF2CC")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

    def style_row(ws, row, cols, fill=None):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            if fill:
                cell.fill = fill
            cell.border = thin_border

    # ── シート1: 収入 ──
    ws1 = wb.active
    ws1.title = "収入"

    headers1 = ["日付", "クライアント名", "案件名", "請求金額(税抜)",
                "消費税(10%)", "請求金額(税込)", "入金日", "入金確認",
                "請求書番号", "備考"]
    col_widths1 = [12, 20, 25, 18, 14, 18, 12, 10, 14, 20]

    for i, (h, w) in enumerate(zip(headers1, col_widths1), 1):
        ws1.cell(1, i, h)
        ws1.column_dimensions[get_column_letter(i)].width = w
    style_header(ws1, 1, len(headers1))

    # データ行 (2-31行: 30行分)
    for row in range(2, 32):
        ws1.cell(row, 5).value = f"=IF(D{row}=\"\",\"\",D{row}*0.1)"
        ws1.cell(row, 6).value = f"=IF(D{row}=\"\",\"\",D{row}+E{row})"
        for col in [4, 5, 6]:
            ws1.cell(row, col).number_format = '¥#,##0'
        ws1.cell(row, 1).number_format = 'yyyy/mm/dd'
        ws1.cell(row, 7).number_format = 'yyyy/mm/dd'
        style_row(ws1, row, len(headers1))

    # 合計行
    ws1.cell(32, 1, "【月計】")
    ws1.cell(32, 4).value = "=SUM(D2:D31)"
    ws1.cell(32, 5).value = "=SUM(E2:E31)"
    ws1.cell(32, 6).value = "=SUM(F2:F31)"
    for col in [4, 5, 6]:
        ws1.cell(32, col).number_format = '¥#,##0'
    style_row(ws1, 32, len(headers1), fill=total_fill)

    ws1.freeze_panes = "A2"

    # ── シート2: 経費 ──
    ws2 = wb.create_sheet("経費")
    headers2 = ["日付", "科目", "金額", "支払方法", "摘要", "レシートあり"]
    col_widths2 = [12, 16, 14, 14, 30, 12]

    for i, (h, w) in enumerate(zip(headers2, col_widths2), 1):
        ws2.cell(1, i, h)
        ws2.column_dimensions[get_column_letter(i)].width = w
    style_header(ws2, 1, len(headers2))

    # 科目プルダウン
    category_list = "通信費,交通費,ソフトウェア,書籍・研修,外注費,消耗品,接待交際費,その他"
    dv_cat = DataValidation(type="list", formula1=f'"{category_list}"', showErrorMessage=True)
    ws2.add_data_validation(dv_cat)

    # 支払方法プルダウン
    pay_list = "現金,クレジットカード,電子マネー,銀行振込,その他"
    dv_pay = DataValidation(type="list", formula1=f'"{pay_list}"', showErrorMessage=True)
    ws2.add_data_validation(dv_pay)

    for row in range(2, 32):
        ws2.cell(row, 3).number_format = '¥#,##0'
        ws2.cell(row, 1).number_format = 'yyyy/mm/dd'
        dv_cat.sqref += f"B{row}"
        dv_pay.sqref += f"D{row}"
        style_row(ws2, row, len(headers2))

    # 科目別集計（H列以降）
    ws2.cell(1, 8, "科目")
    ws2.cell(1, 9, "合計")
    style_header(ws2, 1, 9)

    categories = ["通信費", "交通費", "ソフトウェア", "書籍・研修", "外注費", "消耗品", "接待交際費", "その他"]
    for i, cat in enumerate(categories, 2):
        ws2.cell(i, 8, cat)
        ws2.cell(i, 9).value = f'=SUMIF(B:B,"{cat}",C:C)'
        ws2.cell(i, 9).number_format = '¥#,##0'
        style_row(ws2, i, 9)

    ws2.cell(len(categories) + 2, 8, "合計")
    ws2.cell(len(categories) + 2, 9).value = f"=SUM(I2:I{len(categories)+1})"
    ws2.cell(len(categories) + 2, 9).number_format = '¥#,##0'
    style_row(ws2, len(categories) + 2, 9, fill=total_fill)
    ws2.column_dimensions["H"].width = 16
    ws2.column_dimensions["I"].width = 14

    ws2.freeze_panes = "A2"

    # ── シート3: 案件管理 ──
    ws3 = wb.create_sheet("案件管理")
    headers3 = ["案件名", "クライアント", "単価(円)", "稼働時間(h)",
                "時給(自動)", "開始日", "終了日", "ステータス", "備考"]
    col_widths3 = [25, 20, 14, 14, 14, 12, 12, 14, 25]

    for i, (h, w) in enumerate(zip(headers3, col_widths3), 1):
        ws3.cell(1, i, h)
        ws3.column_dimensions[get_column_letter(i)].width = w
    style_header(ws3, 1, len(headers3))

    status_list = "提案中,進行中,納品済,請求済,入金済,完了"
    dv_st = DataValidation(type="list", formula1=f'"{status_list}"', showErrorMessage=True)
    ws3.add_data_validation(dv_st)

    for row in range(2, 32):
        ws3.cell(row, 5).value = f"=IF(OR(C{row}=\"\",D{row}=0),\"\",C{row}/D{row})"
        for col in [3, 5]:
            ws3.cell(row, col).number_format = '¥#,##0'
        ws3.cell(row, 6).number_format = 'yyyy/mm/dd'
        ws3.cell(row, 7).number_format = 'yyyy/mm/dd'
        dv_st.sqref += f"H{row}"
        style_row(ws3, row, len(headers3))

    ws3.freeze_panes = "A2"

    # ── シート4: サマリー ──
    ws4 = wb.create_sheet("年間サマリー")

    ws4.cell(1, 1, f"{datetime.now().year}年 収支サマリー")
    ws4.cell(1, 1).font = Font(bold=True, size=14)
    ws4.merge_cells("A1:E1")

    # 月別テーブル
    row_header = ["月", "収入(税抜)", "経費", "手取り", "前月比"]
    for i, h in enumerate(row_header, 1):
        ws4.cell(3, i, h)
    style_header(ws4, 3, 5)

    months = ["1月","2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月"]
    yr = f"YEAR(TODAY())"
    for idx, month_name in enumerate(months, 1):
        r = idx + 3
        ws4.cell(r, 1, month_name)
        # 収入: 収入シートから月別集計
        ws4.cell(r, 2).value = (
            f"=SUMPRODUCT((MONTH(収入!A$2:A$200)={idx})"
            f"*(YEAR(収入!A$2:A$200)={yr})*収入!D$2:D$200)"
        )
        # 経費: 経費シートから月別集計
        ws4.cell(r, 3).value = (
            f"=SUMPRODUCT((MONTH(経費!A$2:A$200)={idx})"
            f"*(YEAR(経費!A$2:A$200)={yr})*経費!C$2:C$200)"
        )
        ws4.cell(r, 4).value = f"=B{r}-C{r}"
        ws4.cell(r, 5).value = f"=IF(ROW()-4<=1,\"\",D{r}-D{r-1})"
        for col in [2, 3, 4, 5]:
            ws4.cell(r, col).number_format = '¥#,##0'
        ws4.cell(r, 1).alignment = center
        style_row(ws4, r, 5)

    # 年間合計
    total_row = 16
    ws4.cell(total_row, 1, "年間合計")
    for col, col_letter in [(2, "B"), (3, "C"), (4, "D")]:
        ws4.cell(total_row, col).value = f"=SUM({col_letter}4:{col_letter}15)"
        ws4.cell(total_row, col).number_format = '¥#,##0'
    style_row(ws4, total_row, 5, fill=total_fill)

    # 最高月・最低月
    ws4.cell(18, 1, "最高収入月")
    ws4.cell(18, 2).value = "=TEXT(MATCH(MAX(B4:B15),B4:B15,0),\"0\")&\"月\""
    ws4.cell(18, 3, "最高額")
    ws4.cell(18, 4).value = "=MAX(B4:B15)"
    ws4.cell(18, 4).number_format = '¥#,##0'

    ws4.cell(19, 1, "最低収入月")
    ws4.cell(19, 2).value = "=TEXT(MATCH(MIN(IF(B4:B15>0,B4:B15)),B4:B15,0),\"0\")&\"月\""
    ws4.cell(19, 3, "最低額")
    ws4.cell(19, 4).value = "=MIN(IF(B4:B15>0,B4:B15))"
    ws4.cell(19, 4).number_format = '¥#,##0'

    for col in [1, 2, 3, 4, 5]:
        ws4.column_dimensions[get_column_letter(col)].width = 16

    # 保存
    path = OUTPUT_DIR / "vol1_freelance_cashflow.xlsx"
    wb.save(path)
    print(f"✅ Vol.1 生成: {path}")
    return path


# ─────────────────────────────────────────────────────────
# Vol.2: SNS投稿カレンダー
# ─────────────────────────────────────────────────────────

def create_vol2():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("[ERROR] pip install openpyxl")
        return None

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="2E86AB")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # ── シート1: 月次カレンダー ──
    ws1 = wb.active
    ws1.title = "投稿カレンダー"

    headers = ["日付", "曜日", "媒体", "テーマ・ネタ", "投稿文(下書き)",
               "ハッシュタグ", "画像あり", "投稿済", "いいね数", "備考"]
    col_widths = [12, 8, 18, 25, 45, 30, 10, 10, 10, 20]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws1.cell(1, i, h)
        ws1.column_dimensions[get_column_letter(i)].width = w
    style_header(ws1, 1, len(headers))

    # 媒体プルダウン
    media_list = "Instagram,X(Twitter),Facebook,TikTok,全媒体"
    dv_media = DataValidation(type="list", formula1=f'"{media_list}"')
    ws1.add_data_validation(dv_media)

    for row in range(2, 35):
        ws1.cell(row, 2).value = f'=IF(A{row}="","",TEXT(A{row},"ddd"))'
        ws1.cell(row, 1).number_format = "yyyy/mm/dd"
        dv_media.sqref += f"C{row}"
        for col in range(1, len(headers) + 1):
            ws1.cell(row, col).border = thin_border
        ws1.row_dimensions[row].height = 20

    ws1.freeze_panes = "A2"

    # ── シート2: 投稿テーマ50選 ──
    ws2 = wb.create_sheet("投稿テーマ50選")
    ws2.cell(1, 1, "No")
    ws2.cell(1, 2, "テーマカテゴリ")
    ws2.cell(1, 3, "投稿ネタ")
    ws2.cell(1, 4, "使いやすい媒体")
    style_header(ws2, 1, 4)
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 20

    themes = [
        (1,  "自己紹介",   "自己紹介・プロフィール紹介",         "全媒体"),
        (2,  "サービス",   "提供サービスの概要紹介",              "Instagram/X"),
        (3,  "実績",       "お客様の声・ビフォーアフター",        "Instagram"),
        (4,  "FAQ",        "よくある質問と回答",                  "全媒体"),
        (5,  "豆知識",     "業界の豆知識・役立ち情報",            "全媒体"),
        (6,  "ビフォーアフター", "作業前後の変化を見せる",       "Instagram"),
        (7,  "失敗談",     "失敗したこと・学んだこと",            "X/Facebook"),
        (8,  "日常",       "仕事環境・日常の一コマ",              "Instagram"),
        (9,  "ツール紹介", "使っているおすすめツール",            "全媒体"),
        (10, "季節・行事", "季節のイベントと絡めた投稿",          "全媒体"),
        (11, "数字で見せる","◯◯した結果を数値で公開",             "X/Instagram"),
        (12, "共感",       "フォロワーの悩みに共感する投稿",      "X"),
        (13, "裏話",       "仕事の裏側・制作秘話",                "Instagram"),
        (14, "比較",       "◯◯と△△の違いを比較",               "全媒体"),
        (15, "ノウハウ",   "具体的なHowTo・手順を解説",          "全媒体"),
        (16, "業界ニュース","最新ニュース・トレンドへのコメント", "X"),
        (17, "キャンペーン","期間限定・早割・特典の告知",         "全媒体"),
        (18, "舞台裏",     "仕事中の写真・動画",                  "Instagram"),
        (19, "チーム紹介", "スタッフ・メンバーの紹介",            "Instagram/Facebook"),
        (20, "お客様紹介", "お客様の事例・成功ストーリー",        "全媒体"),
        (21, "本・学び",   "最近読んだ本・得た学び",              "X"),
        (22, "リスト",     "◯◯な人のための△個のヒント",         "全媒体"),
        (23, "問いかけ",   "フォロワーへの質問・アンケート",      "X/Instagram"),
        (24, "記念日",     "開業記念日・節目の報告",              "全媒体"),
        (25, "こだわり",   "仕事のこだわり・価値観",              "全媒体"),
        (26, "目標",       "今月・今年の目標を公開",              "X/Facebook"),
        (27, "振り返り",   "週次・月次の振り返り",                "X"),
        (28, "コラボ",     "他者との対談・コラボ告知",            "全媒体"),
        (29, "受賞・認定", "資格取得・受賞の報告",                "全媒体"),
        (30, "新サービス", "新メニュー・新商品の告知",            "全媒体"),
        (31, "プロセス",   "制作・作業プロセスの公開",            "Instagram"),
        (32, "素材紹介",   "使っている材料・素材のこだわり",      "Instagram"),
        (33, "地域情報",   "地域のイベント・ニュースとの連動",    "Facebook"),
        (34, "感謝",       "フォロワー・お客様への感謝",          "全媒体"),
        (35, "予告",       "次回の投稿・イベントの予告",          "全媒体"),
        (36, "解説",       "専門用語・業界知識の解説",            "全媒体"),
        (37, "トレンド",   "トレンドに乗った投稿",                "X/TikTok"),
        (38, "体験",       "実際に使ってみた・行ってみた",        "Instagram"),
        (39, "インタビュー","お客様・専門家インタビュー",         "YouTube/Facebook"),
        (40, "季節商品",   "季節限定商品・サービスの紹介",        "Instagram"),
        (41, "NG集",       "失敗談・NGシーンを笑いに変える",      "X/TikTok"),
        (42, "まとめ",     "◯◯についてまとめてみた",             "全媒体"),
        (43, "動画ショート","30秒以内のショート動画",             "TikTok/Reels"),
        (44, "投票",       "A vs B どっち派？ の投票",            "X/Instagram"),
        (45, "お得情報",   "割引・特典・無料プレゼント情報",      "全媒体"),
        (46, "読書",       "おすすめ本・読んでよかった本",        "X"),
        (47, "悩み共有",   "フォロワーの悩みを代弁する",          "X"),
        (48, "長所短所",   "サービスの正直なメリット・デメリット","全媒体"),
        (49, "明日から",   "明日から使える◯個のヒント",          "全媒体"),
        (50, "締めの言葉", "月末・週末のまとめ・感謝投稿",        "全媒体"),
    ]

    for no, cat, theme, media in themes:
        ws2.cell(no + 1, 1, no)
        ws2.cell(no + 1, 2, cat)
        ws2.cell(no + 1, 3, theme)
        ws2.cell(no + 1, 4, media)
        for col in range(1, 5):
            ws2.cell(no + 1, col).border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"),
            )
        ws2.cell(no + 1, 1).alignment = Alignment(horizontal="center")

    path = OUTPUT_DIR / "vol2_sns_calendar.xlsx"
    wb.save(path)
    print(f"✅ Vol.2 生成: {path}")
    return path


# ─────────────────────────────────────────────────────────
# メイン
# ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("  販売用テンプレートファイルを生成します")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

    p1 = create_vol1()
    p2 = create_vol2()

    print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"  完了！以下のファイルを BOOTH にアップロードしてください")
    print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    if p1: print(f"  {p1}")
    if p2: print(f"  {p2}")
    print()
