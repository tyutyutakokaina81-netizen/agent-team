"""
05_review.py — 念査（品質チェック）
Claude が成果物を自己レビューし、問題があれば報告・自動修正する。
"""

import csv
import json
import os
import urllib.request
from pathlib import Path

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

REVIEW_PROMPT = """
あなたはデータ品質チェックの専門家です。
以下の成果物データをレビューして、問題点を報告してください。

# 成果物の種類
{file_type}

# データサンプル（先頭20件）
{sample}

# チェック項目
1. 空欄・nullが不自然に多い列はないか
2. 明らかなフォーマットエラー（日付・電話番号・メールなど）
3. 重複データの有無
4. 文字化け・エンコード問題

# 出力形式（JSONのみ）
{{
  "passed": true/false,
  "total_rows": 件数,
  "issues": [
    {{"type": "問題種別", "column": "列名", "detail": "詳細", "severity": "high/medium/low"}}
  ],
  "summary": "全体評価コメント"
}}
"""


def call_claude(prompt: str) -> str:
    if not ANTHROPIC_API_KEY:
        raise EnvironmentError("ANTHROPIC_API_KEY が設定されていません")

    payload = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 1024,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as res:
        body = json.loads(res.read().decode("utf-8"))
    return body["content"][0]["text"]


def _rule_based_review(rows: list[dict], file_type: str) -> dict:
    """API キー不要のルールベース念査（空欄率・重複・文字化け・件数）"""
    import re
    issues = []
    total = len(rows)
    if total == 0:
        return {
            "passed": False,
            "total_rows": 0,
            "issues": [{"type": "empty", "column": "—", "detail": "データが0件です", "severity": "high"}],
            "summary": "成果物が空です",
            "evaluated_by": "rule_based",
        }

    headers = list(rows[0].keys()) if rows else []

    # 1) 空欄率（列ごとに 50% 超なら high、20% 超なら medium）
    for h in headers:
        empties = sum(1 for r in rows if r.get(h) in (None, "", " "))
        rate = empties / total if total else 0
        if rate > 0.5:
            issues.append({"type": "empty", "column": str(h), "detail": f"空欄率 {rate:.0%}", "severity": "high"})
        elif rate > 0.2:
            issues.append({"type": "empty", "column": str(h), "detail": f"空欄率 {rate:.0%}", "severity": "medium"})

    # 2) 重複行
    seen, dup_count = set(), 0
    for r in rows:
        key = tuple((k, str(v)) for k, v in r.items())
        if key in seen:
            dup_count += 1
        else:
            seen.add(key)
    if dup_count > 0:
        sev = "high" if dup_count / total > 0.05 else "medium"
        issues.append({"type": "duplicate", "column": "—", "detail": f"重複 {dup_count} 件", "severity": sev})

    # 3) 文字化け（U+FFFD・連続する ? や ⌘ など）
    mojibake_pat = re.compile(r"[�]|[-]")
    mojibake_rows = sum(1 for r in rows if any(mojibake_pat.search(str(v) or "") for v in r.values()))
    if mojibake_rows > 0:
        issues.append({"type": "encoding", "column": "—", "detail": f"文字化け疑い {mojibake_rows} 行", "severity": "high"})

    # 4) 明らかなフォーマットエラー（メールに@がない等）
    for h in headers:
        if any(k in str(h).lower() for k in ["mail", "メール"]):
            bad = sum(1 for r in rows if r.get(h) and "@" not in str(r[h]))
            if bad > 0:
                issues.append({"type": "format", "column": str(h), "detail": f"@ 欠落 {bad} 件", "severity": "medium"})

    high_count = sum(1 for i in issues if i["severity"] == "high")
    passed = high_count == 0
    return {
        "passed": passed,
        "total_rows": total,
        "issues": issues,
        "summary": (
            f"{file_type} {total}行 / 重大{high_count}件・軽微{len(issues)-high_count}件"
            if issues else f"{file_type} {total}行・問題なし"
        ),
        "evaluated_by": "rule_based",
    }


def _api_review(rows: list[dict], file_type: str) -> dict:
    sample = json.dumps(rows[:20], ensure_ascii=False, indent=2, default=str)
    prompt = REVIEW_PROMPT.format(file_type=file_type, sample=sample)
    raw = call_claude(prompt)
    start = raw.find("{")
    end = raw.rfind("}") + 1
    result = json.loads(raw[start:end])
    result["total_rows"] = len(rows)
    return result


def review_csv(file_path: Path) -> dict:
    rows = []
    with open(file_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    if ANTHROPIC_API_KEY:
        try:
            return _api_review(rows, "CSV")
        except Exception:
            pass  # API 失敗時はルールベースにフォールバック
    return _rule_based_review(rows, "CSV")


def review_excel(file_path: Path) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl が必要です")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    if ANTHROPIC_API_KEY:
        try:
            return _api_review(rows, "Excel")
        except Exception:
            pass
    return _rule_based_review(rows, "Excel")


def run(file_path: str | Path) -> dict:
    file_path = Path(file_path)
    print(f"[念査開始] {file_path.name}")

    if file_path.suffix == ".csv":
        result = review_csv(file_path)
    elif file_path.suffix in (".xlsx", ".xls"):
        result = review_excel(file_path)
    else:
        result = {"passed": False, "error": f"未対応のファイル形式: {file_path.suffix}"}

    # 結果表示
    passed = result.get("passed", False)
    status = "✅ PASS" if passed else "❌ FAIL"
    print(f"[念査結果] {status} | {result.get('total_rows', '?')}行")

    issues = result.get("issues", [])
    if issues:
        print("[問題点]")
        for issue in issues:
            sev = issue.get("severity", "")
            print(f"  [{sev.upper()}] {issue.get('column', '')} - {issue.get('detail', '')}")

    print(f"[評価] {result.get('summary', '')}")
    return result


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        run(sys.argv[1])
    else:
        print("使用方法: python 05_review.py <ファイルパス>")
