"""
04_execute.py — 作業実行
エクセルデータ入力 または Webスクレイピングをカテゴリに応じて実行する。
"""

import csv
import json
import os
import urllib.request
from datetime import datetime
from pathlib import Path

OUTPUT_DIR = Path(__file__).parent.parent / "outputs"
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")


# ─────────────────────────────────────────────
# Excel入力系
# ─────────────────────────────────────────────

def execute_excel_input(job: dict, source_data: list[dict], template_path: str) -> Path:
    """
    source_data: 入力するデータのリスト（[{"col_a": "val", ...}, ...]）
    template_path: ベースとなるExcelファイルのパス
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl が必要です")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ヘッダー行を1行目と仮定してデータを書き込む
    headers = [cell.value for cell in ws[1]]
    for row_idx, record in enumerate(source_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))

    out_path = OUTPUT_DIR / f"{datetime.now().strftime('%Y-%m-%d_%H%M')}_result.xlsx"
    wb.save(out_path)
    print(f"[Excel入力完了] {len(source_data)}行 → {out_path}")
    return out_path


# ─────────────────────────────────────────────
# Webスクレイピング系
# ─────────────────────────────────────────────

def generate_scraping_script(job: dict) -> str:
    """Claude に案件情報を渡してスクレイピングスクリプトを生成させる"""
    if not ANTHROPIC_API_KEY:
        raise EnvironmentError("ANTHROPIC_API_KEY が設定されていません")

    prompt = f"""
以下のクラウドソーシング案件の要件に合わせたPythonスクレイピングスクリプトを生成してください。

# 案件タイトル
{job.get("title", "")}

# 要件
- urllib と html.parser のみ使用（外部ライブラリ不使用）
- 結果をCSV（utf-8-sig）で output.csv に保存
- エラーハンドリング含む
- main() 関数を定義してスクリプト末尾で呼び出す

Pythonコードのみ出力してください（説明不要）。
"""
    payload = json.dumps({
        "model": "claude-sonnet-4-6",
        "max_tokens": 2048,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as res:
        body = json.loads(res.read().decode("utf-8"))
    return body["content"][0]["text"]


SCRAPING_SKELETON = '''"""
スクレイピングスケルトン（手動カスタマイズ用）
案件: {title}

以下の TODO 部分を案件の要件に合わせて書き換えて実行してください。
"""
import csv
import urllib.request
from html.parser import HTMLParser


class SimpleParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.records = []
        # TODO: 案件に応じて状態変数を追加
        # 例) self.in_title = False; self.current = {{}}

    def handle_starttag(self, tag, attrs):
        # TODO: 取得したい要素の開始タグを判定
        pass

    def handle_endtag(self, tag):
        # TODO: 取得した値を self.records に append
        pass

    def handle_data(self, data):
        # TODO: タグ内のテキストを current に格納
        pass


def fetch(url: str) -> str:
    req = urllib.request.Request(
        url,
        headers={{"User-Agent": "Mozilla/5.0"}},
    )
    with urllib.request.urlopen(req, timeout=30) as res:
        return res.read().decode("utf-8", errors="replace")


def main():
    # TODO: 対象 URL を設定
    urls = [
        # "https://example.com/page1",
    ]
    parser = SimpleParser()
    for u in urls:
        parser.feed(fetch(u))

    # CSV 出力（utf-8-sig で Excel 互換）
    if parser.records:
        keys = list(parser.records[0].keys())
        with open("output.csv", "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(parser.records)
        print(f"[完了] {{len(parser.records)}} 件 → output.csv")
    else:
        print("[警告] 取得データが0件です。セレクタ実装を見直してください。")


if __name__ == "__main__":
    main()
'''


def execute_scraping(job: dict) -> Path:
    """スクレイピングスクリプトを生成して実行する。

    API キーがあれば Claude が案件専用スクリプトを生成・実行する。
    API キーがなければ、手動カスタマイズ用スケルトンを保存して人手作業に切替。
    """
    import subprocess
    import tempfile

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")

    # API キーがない場合：スケルトンを保存して人手作業に切替
    if not ANTHROPIC_API_KEY:
        skeleton_path = OUTPUT_DIR / f"{timestamp}_scraping_skeleton.py"
        skeleton_path.write_text(
            SCRAPING_SKELETON.format(title=job.get("title", "")),
            encoding="utf-8",
        )
        print(f"[手動切替] API 未使用のためスケルトンを生成しました")
        print(f"  ファイル: {skeleton_path}")
        print(f"  TODO 部分を埋めて `python3 {skeleton_path.name}` を実行してください")
        return skeleton_path

    # API キーがある場合：Claude にスクリプト生成を依頼
    script_code = generate_scraping_script(job)

    # コードブロック除去
    if "```python" in script_code:
        script_code = script_code.split("```python")[1].split("```")[0].strip()
    elif "```" in script_code:
        script_code = script_code.split("```")[1].split("```")[0].strip()

    # 一時ファイルに保存して実行
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".py", delete=False, encoding="utf-8"
    ) as f:
        f.write(script_code)
        tmp_path = f.name

    out_csv = OUTPUT_DIR / f"{timestamp}_scraping.csv"
    result = subprocess.run(
        ["python3", tmp_path],
        capture_output=True, text=True, timeout=120,
        cwd=str(OUTPUT_DIR),
    )
    Path(tmp_path).unlink(missing_ok=True)

    if result.returncode != 0:
        raise RuntimeError(f"スクレイピング失敗:\n{result.stderr}")

    # output.csv を最終ファイル名にリネーム
    tmp_csv = OUTPUT_DIR / "output.csv"
    if tmp_csv.exists():
        tmp_csv.rename(out_csv)

    print(f"[スクレイピング完了] → {out_csv}")
    return out_csv


def run(job: dict, **kwargs):
    category = job.get("category", "")
    if category == "excel_input":
        source_data = kwargs.get("source_data", [])
        template_path = kwargs.get("template_path", "")
        return execute_excel_input(job, source_data, template_path)
    elif category == "scraping":
        return execute_scraping(job)
    else:
        print(f"[SKIP] 未対応カテゴリ: {category}")
        return None


if __name__ == "__main__":
    # テスト用サンプル
    sample_job = {
        "title": "テスト：サンプルサイトから商品名と価格を収集",
        "category": "scraping",
        "url": "",
    }
    run(sample_job)
