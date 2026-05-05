"""build_template.py — Vol.N 汎用テンプレタ（YAML→ .xlsx + 販売ページmd + publishタスク）

設計書: CDO/outputs/2026-05-05_完全自動化パイプライン設計.md（A1）

使い方:
  python3 scripts/build_template.py <config.yaml>
  → projects/2026-04-08_月30万自動化/C_テンプレ販売/ 配下に
    - Vol{N}_<title>.xlsx
    - vol{N}_listing.md（note/BOOTH 販売ページ）
    - vol{N}_launch_pack.md（ローンチパック）
    - vol{N}_publish_tasks.json（publish.py に追記する用のタスクスニペット）

依存: openpyxl, PyYAML（pyyaml 不在時は stdlib の json/簡易パーサーで代替）
"""
from __future__ import annotations

import datetime as dt
import json
import pathlib
import re
import sys

REPO = pathlib.Path(__file__).resolve().parent.parent
OUT_DIR = REPO / "projects/2026-04-08_月30万自動化/C_テンプレ販売"


def parse_yaml_simple(text: str) -> dict:
    """簡易YAMLパーサー（PyYAML不在環境の最小実装）。
    対応: フラットなキー値、ネストdict、リスト、文字列・数値・bool。
    """
    try:
        import yaml  # type: ignore
        return yaml.safe_load(text)
    except ImportError:
        pass
    # 簡易フォールバック実装
    result = {}
    stack = [(0, result)]
    for raw in text.splitlines():
        if not raw.strip() or raw.lstrip().startswith("#"):
            continue
        indent = len(raw) - len(raw.lstrip())
        line = raw.strip()
        while stack and indent < stack[-1][0]:
            stack.pop()
        ctx = stack[-1][1]
        if line.startswith("- "):
            value = _parse_value(line[2:].strip())
            if isinstance(ctx, list):
                ctx.append(value)
            continue
        if ":" in line:
            key, _, val = line.partition(":")
            key = key.strip()
            val = val.strip()
            if not val:
                # 次行が `- `なら list、そうでなければ dict
                ctx[key] = []
                stack.append((indent + 2, ctx[key]))
            else:
                ctx[key] = _parse_value(val)
    return result


def _parse_value(s: str):
    s = s.strip()
    if s.startswith('"') and s.endswith('"'):
        return s[1:-1]
    if s.startswith("'") and s.endswith("'"):
        return s[1:-1]
    if s.startswith("[") and s.endswith("]"):
        inner = s[1:-1].strip()
        if not inner:
            return []
        return [_parse_value(p.strip()) for p in inner.split(",")]
    if s.lower() == "true":
        return True
    if s.lower() == "false":
        return False
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def slugify(s: str) -> str:
    """ファイル名向けに記号除去・空白アンダースコア化。"""
    s = re.sub(r"[\s\u3000]+", "_", s)
    return re.sub(r"[^\w_\-ぁ-んァ-ン一-龥]", "", s)


def build_xlsx(config: dict, out_path: pathlib.Path) -> int:
    """openpyxl で xlsx を生成。シート定義は config['sheets']。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="4285F4")
    header_align = Alignment(horizontal="center", vertical="center")

    formula_count = 0
    for sheet_def in config.get("sheets", []):
        ws = wb.create_sheet(sheet_def["name"])
        cols = sheet_def.get("columns", [])
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        # データ行（空行 + オプションの数式）
        rows = sheet_def.get("rows", 30)
        formulas = sheet_def.get("formulas", {})  # {"B2": "=A2*2", ...}
        for r in range(2, rows + 2):
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).font = Font(name="Arial", size=11)
        for cell_ref, formula in formulas.items():
            ws[cell_ref] = formula
            formula_count += 1
        # 列幅
        widths = sheet_def.get("widths", {})
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        # データテーブルが空の場合は使い方シート用の追記
        for entry in sheet_def.get("static_rows", []):
            ws.append(entry)

    # 使い方シート
    if config.get("manual"):
        s = wb.create_sheet("使い方", 0)
        for i, line in enumerate(config["manual"], start=1):
            cell = s.cell(row=i, column=1, value=line)
            cell.font = Font(name="Arial", size=11, bold=line.startswith("📘") or line.startswith("📅"))
        s.column_dimensions["A"].width = 70

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return formula_count


def build_listing_md(config: dict, out_path: pathlib.Path) -> None:
    """note/BOOTH 販売ページ md を生成（既存パターン踏襲）。"""
    vol = config["vol"]
    title = config["title"]
    price = config["price"]
    sales = config.get("sales_copy", {})
    pain = sales.get("pain_points", [])
    features = sales.get("features", [])
    target = sales.get("target_audience", "")

    lines = [
        f"# Vol.{vol} {title}",
        "",
        "## note 販売ページ",
        "",
        "### タイトル",
        "```",
        f"{title}",
        "```",
        "",
        "### 本文（noteにそのまま貼る）",
        "```",
    ]
    if pain:
        lines.extend([f'「{p}」' for p in pain])
        lines.append("")
        lines.append("こんな悩みを、このテンプレが解決します。")
        lines.append("")
    if features:
        lines.append("---")
        lines.append("")
        lines.append("■ できること")
        lines.append("")
        for f in features:
            lines.append(f"✅ {f}")
        lines.append("")
    if target:
        lines.append("---")
        lines.append("")
        lines.append("■ こんな方におすすめ")
        lines.append("")
        lines.append(f"・{target}")
        lines.append("")
    lines.extend([
        "---",
        "",
        f"■ 価格：{price}円（税込）",
        "",
        "購入後すぐ使えます。",
        "```",
        "",
        "## BOOTH 掲載用",
        "",
        "### タイトル",
        "```",
        f"{title}",
        "```",
        "",
        "### 説明文",
        "```",
        title,
        "",
        "【含まれるもの】",
    ])
    for sheet_def in config.get("sheets", []):
        lines.append(f"・{sheet_def['name']}")
    lines.extend([
        "",
        "【動作環境】",
        "・Googleスプレッドシート（無料）",
        "・スマホ・PC両対応",
        "",
        "【返金について】",
        "デジタルコンテンツの性質上、返金はお受けできません。",
        "```",
        "",
        f"### 価格：{price}円",
    ])
    out_path.write_text("\n".join(lines), encoding="utf-8")


def build_publish_tasks(config: dict, out_path: pathlib.Path) -> None:
    """publish.py に追記するタスクスニペットをJSONで出力。"""
    vol = config["vol"]
    title = config["title"]
    next_id = 11  # publish.py の既存タスクが10件のため
    tasks = [
        {
            "id": next_id, "name": f"Vol.{vol} note タイトル", "url": "https://note.com/notes/new",
            "content": title,
            "hint": f"→ 新規記事 → 有料記事 → タイトル欄 → 価格{config['price']}円 → タグ設定 → 公開",
        },
        {
            "id": next_id + 1, "name": f"Vol.{vol} BOOTH タイトル", "url": "https://manage.booth.pm/items/new",
            "content": title,
            "hint": f"→ アイテム情報のタイトル欄 → 説明文は vol{vol}_listing.md の BOOTH説明文を別途貼付",
        },
    ]
    out_path.write_text(json.dumps(tasks, ensure_ascii=False, indent=2), encoding="utf-8")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__, file=sys.stderr)
        return 2
    config_path = pathlib.Path(argv[1]).resolve()
    if not config_path.exists():
        print(f"❌ config not found: {config_path}", file=sys.stderr)
        return 2

    config = parse_yaml_simple(config_path.read_text(encoding="utf-8"))
    vol = config.get("vol")
    title = config.get("title", "")
    if not vol or not title:
        print("❌ config に vol と title は必須", file=sys.stderr)
        return 2

    base = OUT_DIR / f"Vol{vol}_{slugify(title)}"
    xlsx_path = base.with_suffix(".xlsx")
    listing_md = OUT_DIR / f"vol{vol}_listing.md"
    publish_json = OUT_DIR / f"vol{vol}_publish_tasks.json"

    formula_count = build_xlsx(config, xlsx_path)
    build_listing_md(config, listing_md)
    build_publish_tasks(config, publish_json)

    print(json.dumps({
        "status": "success",
        "vol": vol,
        "outputs": {
            "xlsx": str(xlsx_path),
            "listing_md": str(listing_md),
            "publish_tasks_json": str(publish_json),
        },
        "formula_count": formula_count,
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
